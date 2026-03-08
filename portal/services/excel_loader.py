"""Reads the investigation Excel and loads companies into state."""
import re
import openpyxl
from pathlib import Path


def slug(name: str) -> str:
    return re.sub(r"[^\w]", "_", str(name).lower())[:40]


def load_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Find best sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "INVEST" in name.upper() or "ALTA" in name.upper():
            target_sheet = wb[name]
            break
    if not target_sheet:
        target_sheet = wb.active

    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Map headers
    headers = [str(c or "").upper().strip() for c in rows[0]]

    def col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return -1

    idx = {
        "num":       col("#"),
        "empresa":   col("EMPRESA", "NOMBRE"),
        "sector":    col("SECTOR"),
        "municipio": col("MUNICIPIO"),
        "empleados": col("EMPLEADOS"),
        "tel":       col("TEL", "TELEFONO"),
        "todos_tel": col("TODOS", "TELEFONOS"),
        "emails":    col("EMAIL"),
        "url":       col("SITIO", "WEB", "URL"),
        "calidad":   col("CALIDAD"),
        "chat_ia":   col("CHAT"),
        "whatsapp":  col("WHATSAPP"),
        "target":    col("TARGET"),
        "facebook":  col("FACEBOOK"),
        "instagram": col("INSTAGRAM"),
        "linkedin":  col("LINKEDIN"),
        "descripcion": col("DESCRIPCION"),
        "servicios": col("SERVICIOS"),
        "propuesta": col("PROPUESTA"),
        "prioridad": col("PRIORIDAD"),
        "salarios":  col("SALARIO"),
    }

    def get(row, key):
        i = idx.get(key, -1)
        if i < 0 or i >= len(row):
            return ""
        v = row[i]
        return str(v).strip() if v is not None else ""

    companies = []
    for row in rows[1:]:
        nombre = get(row, "empresa")
        if not nombre or nombre.upper() in ("EMPRESA", "NOMBRE", ""):
            continue

        c = {
            "id":          slug(nombre),
            "nombre":      nombre,
            "sector":      get(row, "sector"),
            "municipio":   get(row, "municipio"),
            "empleados":   int(get(row, "empleados") or 0),
            "telefono":    get(row, "tel"),
            "telefonos":   get(row, "todos_tel") or get(row, "tel"),
            "emails":      get(row, "emails"),
            "url":         get(row, "url"),
            "calidad_web": get(row, "calidad"),
            "chat_ia":     get(row, "chat_ia"),
            "whatsapp":    get(row, "whatsapp"),
            "es_target":   get(row, "target"),
            "facebook":    get(row, "facebook"),
            "instagram":   get(row, "instagram"),
            "linkedin":    get(row, "linkedin"),
            "descripcion": get(row, "descripcion"),
            "servicios":   get(row, "servicios"),
            "propuesta":   get(row, "propuesta"),
            "prioridad":   get(row, "prioridad"),
            "salarios":    int(float(get(row, "salarios") or 0)),
            "researched":  False,
            "package_ready": False,
        }
        companies.append(c)

    wb.close()
    return companies
