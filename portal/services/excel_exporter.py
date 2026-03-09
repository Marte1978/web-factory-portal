"""
Generador de Excel completo para Web Factory Portal.
Exporta toda la investigación OSINT de empresas dominicanas.
"""
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation


# ─── Colores de marca ────────────────────────────────────────────────────────

C_AZUL_OSCURO  = "1E3A5F"
C_AZUL_MEDIO   = "2B5EA7"
C_AZUL_CLARO   = "E8F0FE"
C_VERDE        = "27AE60"
C_AMARILLO     = "F39C12"
C_ROJO         = "E74C3C"
C_GRIS_CLARO   = "F8F9FA"
C_GRIS_BORDE   = "DEE2E6"
C_BLANCO       = "FFFFFF"
C_NARANJA      = "E67E22"


def _header_style(ws, row, col, value, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(color=C_BLANCO, bold=True, size=10, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=C_AZUL_OSCURO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _data_cell(ws, row, col, value, fill=None, font_color=None, bold=False, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, color=font_color or "000000", bold=bold)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    return cell


def _thin_border():
    thin = Side(style="thin", color=C_GRIS_BORDE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_companies_excel(companies: List[Dict], output_path: Path) -> Path:
    """
    Genera reporte Excel completo con 4 hojas:
    1. Prospectos — todos los datos OSINT
    2. Targets — solo empresas con es_target=SI
    3. Con Maps — empresas con rating de Google Maps
    4. Resumen — estadísticas y gráfico
    """
    wb = openpyxl.Workbook()

    # ── HOJA 1: TODOS LOS PROSPECTOS ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Todos los Prospectos"

    columns = [
        # (header, key_en_dict, width)
        # IDENTIFICACIÓN
        ("Empresa",          "nombre",         38),
        ("Sector",           "sector",         20),
        ("Ciudad/Municipio", "municipio",       20),
        ("RNC",              "rnc",             14),
        ("Empleados",        "empleados",       10),
        # WEB
        ("Sitio Web",        "url",             32),
        ("Calidad Web",      "calidad_web",     14),
        ("Score (0-8)",      "score_web",       12),
        ("Título Web",       "titulo_web",      30),
        ("Chat IA",          "chat_ia",         10),
        ("WhatsApp Web",     "whatsapp",        12),
        # CONTACTO
        ("Teléfonos",        "telefonos",       22),
        ("Emails",           "emails",          28),
        ("Dirección",        "direccion",       35),
        # GOOGLE MAPS
        ("Maps Rating",      "gmaps_rating",    12),
        ("Maps Reseñas",     "gmaps_reviews",   13),
        ("Horarios",         "gmaps_hours",     30),
        ("Link Maps",        "gmaps_url",       25),
        # REDES SOCIALES
        ("Facebook",         "facebook",        28),
        ("Instagram",        "instagram",       28),
        ("LinkedIn",         "linkedin",        28),
        ("YouTube",          "youtube",         25),
        ("TikTok",           "tiktok",          20),
        ("WhatsApp Link",    "whatsapp_link",   20),
        # ANÁLISIS
        ("Es Target",        "es_target",       10),
        ("Propuesta de Valor","propuesta",      50),
        # META
        ("Investigado",      "researched",      12),
        ("Fecha Invest.",     "research_date",  16),
    ]

    # Grupos de columnas para colores de encabezado alternados
    col_groups = {
        "IDENTIFICACIÓN":  (1, 5,  "1E3A5F"),
        "PRESENCIA WEB":   (6, 11, "1A5276"),
        "CONTACTO":        (12, 14,"154360"),
        "GOOGLE MAPS":     (15, 18,"1B4F72"),
        "REDES SOCIALES":  (19, 24,"0E6655"),
        "ANÁLISIS":        (25, 26,"6E2F1A"),
        "META":            (27, 28,"1C2833"),
    }

    # Fila 1: grupo labels
    ws.row_dimensions[1].height = 18
    for group_name, (c_start, c_end, color) in col_groups.items():
        ws.merge_cells(start_row=1, start_column=c_start,
                      end_row=1, end_column=c_end)
        cell = ws.cell(row=1, column=c_start, value=group_name)
        cell.font = Font(color=C_BLANCO, bold=True, size=9, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 2: headers individuales
    ws.row_dimensions[2].height = 28
    for col_idx, (header, key, width) in enumerate(columns, 1):
        _header_style(ws, 2, col_idx, header, width)

    # Freeze panes (2 filas de encabezado + columna nombre)
    ws.freeze_panes = "B3"

    # Datos
    for row_idx, co in enumerate(companies, 3):
        bg = C_GRIS_CLARO if row_idx % 2 == 0 else C_BLANCO
        ws.row_dimensions[row_idx].height = 16

        for col_idx, (header, key, width) in enumerate(columns, 1):
            value = co.get(key, "") or ""
            font_color = None
            bold = False
            center = False

            # Coloreado especial por columna
            if key == "calidad_web":
                color_map = {"PROFESIONAL": C_VERDE, "INTERMEDIA": C_AMARILLO,
                             "BASICA": C_ROJO, "SIN WEB": "888888"}
                font_color = color_map.get(str(value), None)
                bold = bool(value)
                center = True
            elif key == "es_target":
                font_color = C_VERDE if value == "SI" else (C_ROJO if value == "NO" else None)
                bold = bool(value)
                center = True
            elif key == "chat_ia" or key == "whatsapp":
                font_color = C_VERDE if value == "SI" else (C_ROJO if value == "NO" else None)
                center = True
            elif key == "gmaps_rating" and value:
                try:
                    r_val = float(value)
                    font_color = C_VERDE if r_val >= 4.0 else (C_AMARILLO if r_val >= 3.0 else C_ROJO)
                    bold = True
                    center = True
                except Exception:
                    pass
            elif key == "score_web" and value != "":
                try:
                    s_val = int(value)
                    font_color = C_VERDE if s_val >= 6 else (C_AMARILLO if s_val >= 4 else C_ROJO)
                    bold = True
                    center = True
                except Exception:
                    pass
            elif key in ("url", "gmaps_url", "facebook", "instagram", "linkedin",
                         "youtube", "tiktok", "whatsapp_link") and value:
                font_color = C_AZUL_MEDIO

            cell = _data_cell(ws, row_idx, col_idx, value, fill=bg,
                              font_color=font_color, bold=bold, center=center)
            cell.border = _thin_border()

            # Hyperlinks para URLs
            if key in ("url", "gmaps_url", "facebook", "instagram", "linkedin",
                       "youtube", "tiktok") and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
                cell.font = Font(color=C_AZUL_MEDIO, underline="single",
                                 size=9, name="Calibri")

    ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{len(companies)+2}"

    # ── HOJA 2: SOLO TARGETS ─────────────────────────────────────────────────
    targets = [c for c in companies if c.get("es_target") == "SI"]
    ws2 = wb.create_sheet("Targets (Prospectos)")
    _build_targets_sheet(ws2, targets)

    # ── HOJA 3: CON MAPS RATING ───────────────────────────────────────────────
    con_maps = sorted(
        [c for c in companies if c.get("gmaps_rating")],
        key=lambda x: float(x.get("gmaps_rating") or 0), reverse=True
    )
    ws3 = wb.create_sheet("Con Rating Google Maps")
    _build_maps_sheet(ws3, con_maps)

    # ── HOJA 4: RESUMEN ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumen y Estadísticas")
    _build_summary_sheet(ws4, companies, targets)

    wb.save(str(output_path))
    return output_path


def _build_targets_sheet(ws, companies: List[Dict]):
    """Hoja de prospectos target — las más importantes para ventas."""
    ws.row_dimensions[1].height = 30

    cols = [
        ("Empresa", "nombre", 38),
        ("Sector", "sector", 20),
        ("Ciudad", "municipio", 18),
        ("Web Actual", "url", 30),
        ("Calidad", "calidad_web", 13),
        ("Maps Rating", "gmaps_rating", 12),
        ("Reseñas", "gmaps_reviews", 10),
        ("Teléfonos", "telefonos", 22),
        ("Emails", "emails", 28),
        ("Facebook", "facebook", 25),
        ("Instagram", "instagram", 25),
        ("Le falta", "propuesta", 50),
        ("Horarios", "gmaps_hours", 25),
    ]

    for col_idx, (header, key, width) in enumerate(cols, 1):
        _header_style(ws, 1, col_idx, header, width)

    ws.freeze_panes = "A2"

    for row_idx, co in enumerate(companies, 2):
        bg = C_GRIS_CLARO if row_idx % 2 == 0 else C_BLANCO
        ws.row_dimensions[row_idx].height = 16
        for col_idx, (header, key, width) in enumerate(cols, 1):
            val = co.get(key, "") or ""
            font_color = None
            bold = False
            if key == "calidad_web":
                color_map = {"PROFESIONAL": C_VERDE, "INTERMEDIA": C_AMARILLO, "BASICA": C_ROJO}
                font_color = color_map.get(str(val))
                bold = True
            cell = _data_cell(ws, row_idx, col_idx, val, fill=bg,
                              font_color=font_color, bold=bold)
            cell.border = _thin_border()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(companies)+1}"


def _build_maps_sheet(ws, companies: List[Dict]):
    """Hoja de empresas con rating de Google Maps."""
    ws.row_dimensions[1].height = 30

    cols = [
        ("Empresa", "nombre", 38),
        ("Ciudad", "municipio", 18),
        ("Rating Maps", "gmaps_rating", 13),
        ("Reseñas", "gmaps_reviews", 12),
        ("Horarios", "gmaps_hours", 35),
        ("Dirección", "direccion", 35),
        ("Teléfono (Maps)", "phone_maps", 20),
        ("Web (Maps)", "website_maps", 30),
        ("Link Maps", "gmaps_url", 28),
    ]

    for col_idx, (header, key, width) in enumerate(cols, 1):
        _header_style(ws, 1, col_idx, header, width)

    ws.freeze_panes = "A2"

    for row_idx, co in enumerate(companies, 2):
        bg = C_GRIS_CLARO if row_idx % 2 == 0 else C_BLANCO
        ws.row_dimensions[row_idx].height = 16
        for col_idx, (header, key, width) in enumerate(cols, 1):
            val = co.get(key, "") or ""
            font_color = None
            bold = False
            center = False
            if key == "gmaps_rating" and val:
                try:
                    r_val = float(val)
                    font_color = C_VERDE if r_val >= 4.0 else (C_AMARILLO if r_val >= 3.0 else C_ROJO)
                    bold = True
                    center = True
                except Exception:
                    pass
            cell = _data_cell(ws, row_idx, col_idx, val, fill=bg,
                              font_color=font_color, bold=bold, center=center)
            cell.border = _thin_border()
            if key == "gmaps_url" and str(val).startswith("http"):
                cell.hyperlink = str(val)


def _build_summary_sheet(ws, all_companies: List[Dict], targets: List[Dict]):
    """Hoja de resumen con estadísticas clave."""
    # Título
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "WEB FACTORY PORTAL — REPORTE DE INVESTIGACIÓN"
    title.font = Font(color=C_BLANCO, bold=True, size=16, name="Calibri")
    title.fill = PatternFill("solid", fgColor=C_AZUL_OSCURO)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | República Dominicana"
    sub.font = Font(color="555555", italic=True, size=10)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    total = len(all_companies)
    n_targets = len(targets)
    con_web = sum(1 for c in all_companies if c.get("url"))
    profesional = sum(1 for c in all_companies if c.get("calidad_web") == "PROFESIONAL")
    intermedia = sum(1 for c in all_companies if c.get("calidad_web") == "INTERMEDIA")
    basica = sum(1 for c in all_companies if c.get("calidad_web") == "BASICA")
    sin_web = total - con_web
    con_maps = sum(1 for c in all_companies if c.get("gmaps_rating"))
    con_chat = sum(1 for c in all_companies if c.get("chat_ia") == "SI")
    con_wa = sum(1 for c in all_companies if c.get("whatsapp") == "SI")
    con_fb = sum(1 for c in all_companies if c.get("facebook"))
    con_ig = sum(1 for c in all_companies if c.get("instagram"))
    investigadas = sum(1 for c in all_companies if c.get("researched"))

    stats = [
        ("", "MÉTRICAS GENERALES", "", "", "", ""),
        ("Total empresas en base", total, "", "Total investigadas", investigadas, ""),
        ("Prospectos TARGET", n_targets, f"{n_targets/total*100:.0f}%" if total else "0%", "Sin investigar", total-investigadas, ""),
        ("", "", "", "", "", ""),
        ("", "PRESENCIA DIGITAL", "", "", "", ""),
        ("Con sitio web", con_web, f"{con_web/total*100:.0f}%" if total else "0%", "Sin sitio web", sin_web, f"{sin_web/total*100:.0f}%" if total else "0%"),
        ("Web PROFESIONAL", profesional, f"{profesional/total*100:.0f}%" if total else "0%", "Web INTERMEDIA", intermedia, f"{intermedia/total*100:.0f}%" if total else "0%"),
        ("Web BASICA", basica, f"{basica/total*100:.0f}%" if total else "0%", "Con Google Maps", con_maps, f"{con_maps/total*100:.0f}%" if total else "0%"),
        ("", "", "", "", "", ""),
        ("", "HERRAMIENTAS DIGITALES", "", "", "", ""),
        ("Con Chat IA", con_chat, f"{con_chat/total*100:.0f}%" if total else "0%", "Con WhatsApp Web", con_wa, f"{con_wa/total*100:.0f}%" if total else "0%"),
        ("Con Facebook", con_fb, f"{con_fb/total*100:.0f}%" if total else "0%", "Con Instagram", con_ig, f"{con_ig/total*100:.0f}%" if total else "0%"),
    ]

    for row_idx, row_data in enumerate(stats, 4):
        ws.row_dimensions[row_idx + 3].height = 18
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 3, column=col_idx, value=val)
            if col_idx in (1, 4) and val and not val.startswith(" "):
                cell.font = Font(bold=True, size=10, name="Calibri")
            elif col_idx in (2, 5) and isinstance(val, int):
                cell.font = Font(color=C_AZUL_MEDIO, bold=True, size=12, name="Calibri")
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (3, 6):
                cell.font = Font(color="666666", size=9, name="Calibri")
                cell.alignment = Alignment(horizontal="center")
            # Section headers
            if col_idx == 2 and isinstance(val, str) and val.isupper():
                ws.merge_cells(start_row=row_idx+3, start_column=1,
                               end_row=row_idx+3, end_column=6)
                cell.font = Font(color=C_BLANCO, bold=True, size=10, name="Calibri")
                cell.fill = PatternFill("solid", fgColor=C_AZUL_MEDIO)
                cell.alignment = Alignment(horizontal="center")
                break

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10


def generate_report(companies: List[Dict], output_dir: Path = None) -> Path:
    """Entry point: genera el Excel y lo guarda."""
    if output_dir is None:
        output_dir = Path("C:/Users/Willy/sistema de egocios/reports")
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"web-factory-reporte-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    output_path = output_dir / filename

    return export_companies_excel(companies, output_path)
