"""
Investigador 100 Empresas - High Profile Targets
Lee directamente del raw data, filtra las mejores locales,
investiga cada una y exporta reporte completo.
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import re, os, time, json, requests, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from bs4 import BeautifulSoup
try:
    from ddgs import DDGS
except ImportError:
    from duckduckgo_search import DDGS
from urllib.parse import urljoin, urlparse
from io import BytesIO
from PIL import Image as PILImage
from datetime import datetime

# ─── Config ───────────────────────────────────────────────────────────────────

RAW_EXCEL   = r"C:\Users\Willy\Downloads\empresas_2.xlsx"
OUTPUT      = r"C:\Users\Willy\OneDrive\Escritorio\INVESTIGACION_100_EMPRESAS.xlsx"
LOGOS_DIR   = r"C:\Users\Willy\sistema de egocios\research\logos"
PROGRESS_F  = r"C:\Users\Willy\sistema de egocios\research\progress_100.json"
N           = 100

os.makedirs(LOGOS_DIR, exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}

# Palabras que indican multinacional o entidad que ya tiene web global
MULTINACIONAL_KW = [
    "colgate", "palmolive", "kpmg", "pricewaterhouse", "deloitte", "ernst&young",
    "ernst & young", "bayer", "kimberly", "mondelez", "mead johnson", "3m dominicana",
    "american airlines", "pan am", "wartsila", "berry plastics", "berry dominicana",
    "payless", "scotiabank", "citibank", "honda", "toyota", "ford ", "hyundai",
    "samsung", "apple ", "microsoft", "oracle ", "ibm ", "xerox", "siemens",
    "nestle", "unilever", "procter", "johnson & johnson", "abbott ", "pfizer",
    "novartis", "roche ", "organismo coordinador", "junta de aviacion",
    "iglesia de jesucristo", "plan republica", "zona franca industrial de las americas",
    "compania de electricidad", "empresa generadora", "transporte de gas",
    "dominican power", "helicopteros dominicanos", "generadora palamara",
    "transporte lpg", "nuts trading", "quimocaribe", "inter-quimica",
    "productos quimicos industriales", "agroindustrial", "compania comercial caribe",
]

# Sectores que realmente compran web + IA
SECTORES_GOLD = {
    8:  "Hoteles/Turismo",
    11: "Inmobiliaria",
    12: "Servicios Empresariales",
    14: "Educacion",
    15: "Salud/Medicina",
    16: "Servicios Sociales/ONG",
    18: "Otros Servicios",
    7:  "Comercio Minorista",
    10: "Finanzas/Seguros",
}
SECTORES_SILVER = {
    3: "Manufactura",
    5: "Construccion",
    6: "Comercio Mayor",
    9: "Transporte/Logistica",
    33: "Manufactura Avanzada",
    41: "Construccion Especializada",
}
SECTORES_EXCLUIR = {1, 2, 4, 13, 17, 25, 26, 34, 36}

MUNICIPIOS_SD = {
    "DISTRITO NACIONAL", "SANTO DOMINGO ESTE",
    "SANTO DOMINGO OESTE", "SANTO DOMINGO NORTE"
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean(t):
    return re.sub(r"\s+", " ", str(t or "")).strip()

def es_multinacional(nombre):
    n = nombre.lower()
    return any(kw in n for kw in MULTINACIONAL_KW)

def phones_from(text):
    found = re.findall(r"\b8\d{2}[\-\.\s]?\d{3}[\-\.\s]?\d{4}\b", text or "")
    found += re.findall(r"\(\d{3}\)[\s\-]?\d{3}[\-\s]\d{4}", text or "")
    seen, result = set(), []
    for p in found:
        p2 = re.sub(r"\D", "", p)
        if p2 not in seen and len(p2) >= 7:
            seen.add(p2)
            result.append(p2)
    return result[:5]

def emails_from(text):
    found = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text or "")
    return list(set(e for e in found if "example" not in e and "@sentry" not in e))[:3]

def socials_from(text):
    s = {}
    pats = {
        "Facebook":  r"facebook\.com/(?!sharer)[A-Za-z0-9._\-]+",
        "Instagram": r"instagram\.com/[A-Za-z0-9._\-]+",
        "LinkedIn":  r"linkedin\.com/(?:company|in)/[A-Za-z0-9._\-]+",
        "Twitter":   r"(?:twitter|x)\.com/[A-Za-z0-9._\-]+",
        "YouTube":   r"youtube\.com/(?:channel|user|@)[A-Za-z0-9._\-/]+",
        "WhatsApp":  r"(?:wa\.me|whatsapp\.com/send)[?/\d]+",
    }
    for red, pat in pats.items():
        m = re.search(pat, text or "", re.IGNORECASE)
        if m:
            link = m.group(0)
            s[red] = link if link.startswith("http") else "https://" + link
    return s

def tiene_chat(html):
    kw = ["tawk.to", "intercom", "zendesk", "drift", "tidio", "crisp.chat",
          "freshchat", "livechat", "chatbot", "asistente virtual", "habla con nosotros",
          "chat en vivo", "chat en tiempo real", "hubspot", "manychat"]
    h = (html or "").lower()
    return any(k in h for k in kw)

def tiene_whatsapp(html):
    return any(k in (html or "").lower() for k in ["wa.me", "api.whatsapp", "whatsapp"])

def score_web(soup, html, url):
    if not url:
        return "SIN WEB", 0
    if not soup:
        return "CAIDA", 0
    s = 1
    if soup.find("meta", {"name": "description"}): s += 1
    if len(soup.find_all("img")) > 3: s += 1
    if soup.find("form"): s += 1
    if soup.find("meta", {"name": "viewport"}): s += 1
    if tiene_chat(html): s += 2
    if tiene_whatsapp(html): s += 1
    labels = {0:"SIN WEB",1:"BASICA",2:"BASICA",3:"INTERMEDIA",
              4:"INTERMEDIA",5:"BUENA",6:"BUENA",7:"PROFESIONAL",8:"PROFESIONAL"}
    return labels.get(min(s,8), "BASICA"), s

def descargar_logo(base_url, nombre, soup):
    if not soup or not base_url:
        return None
    try:
        og = soup.find("meta", property="og:image")
        url = None
        if og and og.get("content"):
            url = urljoin(base_url, og["content"])
        else:
            fav = soup.find("link", rel=lambda r: r and "icon" in " ".join(r if r else []).lower())
            if fav and fav.get("href"):
                url = urljoin(base_url, fav["href"])
            else:
                url = urljoin(base_url, "/favicon.ico")
        r = requests.get(url, timeout=7, headers=HEADERS)
        if r.status_code == 200 and len(r.content) > 400:
            img = PILImage.open(BytesIO(r.content)).convert("RGBA")
            img = img.resize((72, 72), PILImage.LANCZOS)
            slug = re.sub(r"[^\w]", "_", nombre[:20])
            path = os.path.join(LOGOS_DIR, f"{slug}.png")
            img.save(path, "PNG")
            return path
    except Exception:
        pass
    return None

# ─── Busqueda DDG ─────────────────────────────────────────────────────────────

def buscar(nombre):
    nombre_limpio = re.sub(
        r"\b(S\.?\s*A\.?S?|SRL|EIRL|LTD|INC|CORP|S\.?\s*A\.?|CIA|COMPANIA|COMPANY|C\s*POR\s*A|SAS|RL|LDC|LLC)\b",
        "", nombre, flags=re.IGNORECASE).strip()
    nombre_limpio = re.sub(r"\s+", " ", nombre_limpio).strip()

    queries = [
        f'"{nombre_limpio}" santo domingo OR dominicana',
        f'{nombre_limpio} republica dominicana web telefono',
    ]
    results = []
    for q in queries:
        try:
            with DDGS() as d:
                results.extend(d.text(q, max_results=7, region="es-419"))
            time.sleep(0.8)
        except Exception:
            time.sleep(1)
    return results, nombre_limpio

def detectar_url(results, nombre, nombre_limpio):
    BASURA = [
        "facebook.com","instagram.com","linkedin.com","twitter.com","x.com",
        "youtube.com","google.com","wikipedia.org","wikimedia.org",
        "tripadvisor.com","booking.com","trustpilot.com","glassdoor.com",
        "reddit.com","quora.com","amazon.com","ebay.com","mercadolibre.com",
        "zhihu.com","baidu.com","ruwiki.ru","chatgpt.com","openai.com",
        "portcitydaily.com","drudgereport.com","xfinity.com","gostanford.com",
        "wordreference.com","rainews.it","digitei.com","gulfcoast",
        "forums.","forum.","yahoo.com","bing.com","indeed.com",
        "/noticias/","/news/","/articulos/","/wiki/","/forum",
        "sitio-web-rd.com","amarillasinternet","cylex","companias.do",
        "dondeir.com","donde.com.do","1411.com.do","paginas-dominicanas",
        "gob.do","gobierno.gob",
    ]
    stopwords = {
        "sa","sas","srl","eirl","ltd","inc","corp","de","la","el","los","las",
        "del","y","e","the","and","dominicana","dominicano","republica","company",
        "compania","cia","hermanos","interamerica","internacional","group","grupo"
    }
    palabras = [w.lower() for w in re.split(r"[\s\-_&.,]+", nombre_limpio)
                if len(w) >= 4 and w.lower() not in stopwords]

    scored = []
    for r in results:
        url = r.get("href", "")
        title = (r.get("title") or "").lower()
        body  = (r.get("body")  or "").lower()
        if not url:
            continue
        url_l = url.lower()
        if any(b in url_l for b in BASURA):
            continue

        sc = 0
        for pal in palabras:
            if pal in url_l:   sc += 4
            if pal in title:   sc += 2
            if pal in body:    sc += 1

        if ".do/" in url_l or url_l.endswith(".do"): sc += 5
        if any(k in body for k in ["santo domingo","republica dominicana","dominicana"]): sc += 2
        if sc > 0:
            scored.append((sc, url))

    scored.sort(reverse=True)
    return scored[0][1] if scored else ""

def scrape(url):
    if not url:
        return None, "", url
    if not url.startswith("http"):
        url = "https://" + url
    for scheme in ["https://", "http://"]:
        try:
            u = scheme + url.split("://", 1)[-1]
            r = requests.get(u, timeout=12, headers=HEADERS, allow_redirects=True)
            r.encoding = r.apparent_encoding
            return BeautifulSoup(r.text, "lxml"), r.text, r.url
        except Exception:
            pass
    return None, "", url

def extraer(soup, html, url):
    info = {"desc":"","tel":"","email":"","servicios":"","redes":{},"titulo":""}
    if not soup:
        return info
    t = soup.find("title")
    if t: info["titulo"] = clean(t.get_text())[:80]
    m = (soup.find("meta", {"name":"description"}) or
         soup.find("meta", property="og:description"))
    if m: info["desc"] = clean(m.get("content",""))[:300]
    for tag in soup(["script","style","nav","header"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)
    phones = phones_from(text) + phones_from(html)
    info["tel"]   = " | ".join(list(dict.fromkeys(phones))[:3])
    emails = emails_from(text) + emails_from(html)
    info["email"] = " | ".join(list(dict.fromkeys(emails))[:2])
    info["redes"] = socials_from(html)
    kw_svc = ["servicio","producto","ofrecemos","soluciones","especiali",
               "brindamos","nuestros","hacemos","realizamos"]
    svcs = []
    for p in soup.find_all(["p","li","h2","h3"], limit=60):
        t2 = clean(p.get_text())
        if len(t2) > 25 and any(k in t2.lower() for k in kw_svc):
            svcs.append(t2[:110])
        if len(svcs) >= 5: break
    info["servicios"] = " | ".join(svcs)
    return info

def propuesta(nombre, calidad, chat, wa, sector):
    if calidad == "SIN WEB":
        return (f"{nombre} no tiene presencia digital. Lanzamos su sitio profesional con "
                f"chat IA 24/7 y captura por WhatsApp en 48 horas por $X,XXX.")
    elif calidad in ("CAIDA",):
        return (f"El sitio de {nombre} esta caido o inaccesible. "
                f"Podemos reemplazarlo con una web moderna y funcional esta semana.")
    elif calidad in ("BASICA","INTERMEDIA") and not chat:
        return (f"{nombre} tiene web pero sin automatizacion. Un rediseno con IA y "
                f"WhatsApp puede duplicar sus consultas sin contratar personal extra.")
    elif not chat:
        return (f"{nombre} tiene buena web pero pierde leads cada noche sin chat IA. "
                f"Integramos asistente virtual en 3 dias.")
    else:
        return (f"{nombre} tiene presencia digital. Explorar upsell: SEO, "
                f"campanas de WhatsApp o sistema de reservas integrado.")

# ─── Leer top N empresas del raw data ─────────────────────────────────────────

def cargar_empresas():
    wb = openpyxl.load_workbook(RAW_EXCEL, read_only=True, data_only=True)
    ws = wb.active

    sector_map = {**SECTORES_GOLD, **SECTORES_SILVER}
    lista = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rnc, nombre, calle, num, edif, piso, apto, barrio, muni_cod, muni_nom, tipo, tel1, ext1, emp, sal, sec_eco = row
        if str(muni_nom).upper() not in MUNICIPIOS_SD: continue
        if not isinstance(emp, (int, float)) or emp < 15 or emp > 200: continue
        sec = int(sec_eco) if isinstance(sec_eco, (int, float)) else 0
        if sec in SECTORES_EXCLUIR: continue
        if sec not in SECTORES_GOLD and sec not in SECTORES_SILVER: continue

        nombre_s = str(nombre or "").strip()
        if es_multinacional(nombre_s): continue  # filtrar multinationales
        if not nombre_s: continue

        sal_v = sal if isinstance(sal, (int, float)) else 0
        gold  = sec in SECTORES_GOLD
        pri_n = 3 if gold else 2
        score = (emp * pri_n) + (sal_v / 1200)

        tel_s = str(int(tel1)) if isinstance(tel1, (int, float)) else str(tel1 or "")
        dir_p = [str(calle or ""), str(num or ""), str(barrio or "")]
        direc = ", ".join(p for p in dir_p if p and p not in ("None",""))

        lista.append({
            "rnc": rnc, "nombre": nombre_s, "empleados": int(emp),
            "salarios": sal_v, "municipio": str(muni_nom or ""),
            "sector": sector_map.get(sec, f"Sector-{sec}"),
            "prioridad": "ALTA" if gold else "MEDIA",
            "score": score, "telefono": tel_s, "direccion": direc,
        })

    wb.close()
    lista.sort(key=lambda x: x["score"], reverse=True)
    return lista[:N]

# ─── Investigar empresa ───────────────────────────────────────────────────────

def investigar(empresa, idx, total):
    nombre = empresa["nombre"]
    print(f"\n[{idx:3}/{total}] {nombre} ({empresa['empleados']} emp | {empresa['sector']})")

    results, nombre_limpio = buscar(nombre)
    url_oficial = detectar_url(results, nombre, nombre_limpio)
    print(f"         Web: {url_oficial or 'NO ENCONTRADA'}")

    soup, html, url_final = scrape(url_oficial)
    logo = descargar_logo(url_final, nombre, soup) if soup else None
    info = extraer(soup, html, url_final)

    calidad, sc = score_web(soup, html, url_oficial)
    chat = tiene_chat(html)
    wa   = tiene_whatsapp(html)

    # Combinar telefonos
    tels = list(dict.fromkeys(
        ([empresa["telefono"]] if empresa["telefono"] else []) +
        (info["tel"].split(" | ") if info["tel"] else [])
    ))
    emails = list(dict.fromkeys(
        (info["email"].split(" | ") if info["email"] else [])
    ))
    redes = {**socials_from(" ".join(r.get("body","") for r in results)),
             **info["redes"]}

    snippet = ""
    if not info["desc"] and results:
        snippet = clean(results[0].get("body",""))[:280]

    es_target = "NO" if (calidad == "PROFESIONAL" and chat) else "SI"
    print(f"         Calidad: {calidad} | Chat IA: {'SI' if chat else 'NO'} | Target: {es_target}")

    return {
        "nombre": nombre, "rnc": empresa["rnc"],
        "municipio": empresa["municipio"], "sector": empresa["sector"],
        "empleados": empresa["empleados"], "salarios": empresa["salarios"],
        "prioridad": empresa["prioridad"],
        "tel_original": empresa["telefono"],
        "telefonos": " | ".join(t for t in tels if t)[:100],
        "emails": " | ".join(emails)[:100],
        "url": url_final or url_oficial,
        "calidad_web": calidad,
        "score_web": sc,
        "chat_ia": "SI" if chat else "NO",
        "whatsapp": "SI" if wa else "NO",
        "es_target": es_target,
        "facebook": redes.get("Facebook",""),
        "instagram": redes.get("Instagram",""),
        "linkedin": redes.get("LinkedIn",""),
        "twitter": redes.get("Twitter",""),
        "youtube": redes.get("YouTube",""),
        "whatsapp_link": redes.get("WhatsApp",""),
        "descripcion": info["desc"] or snippet,
        "servicios": info["servicios"][:400],
        "titulo_web": info["titulo"],
        "logo": logo,
        "propuesta": propuesta(nombre, calidad, chat, wa, empresa["sector"]),
    }

# ─── Exportar Excel ───────────────────────────────────────────────────────────

COLUMNAS = [
    ("#",2),("LOGO",9),("EMPRESA",38),("PRIORIDAD",10),("SECTOR",20),
    ("MUNICIPIO",16),("EMPLEADOS",10),("TEL ORIGINAL",14),("TODOS TELEFONOS",24),
    ("EMAILS",26),("SITIO WEB",30),("CALIDAD WEB",13),("CHAT IA",9),
    ("WHATSAPP",10),("ES TARGET",10),("FACEBOOK",28),("INSTAGRAM",28),
    ("LINKEDIN",28),("YOUTUBE",28),("WHATSAPP LINK",20),
    ("DESCRIPCION",50),("SERVICIOS",52),("PROPUESTA DE VENTA",60),
]

def exportar(resultados):
    wb = openpyxl.Workbook()

    # ── Tab INVESTIGACION ──
    ws = wb.active
    ws.title = "INVESTIGACION"
    ws.row_dimensions[1].height = 32

    H = PatternFill("solid", fgColor="0D47A1")
    for ci, (col, w) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = H
        c.font = Font(color="FFFFFF", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "C2"

    color_cal = {
        "SIN WEB":"EF5350","CAIDA":"FF7043","BASICA":"FF8A65",
        "INTERMEDIA":"FFD54F","BUENA":"81C784","PROFESIONAL":"42A5F5"
    }
    fills = {
        "SIN WEB":   PatternFill("solid", fgColor="FFEBEE"),
        "CAIDA":     PatternFill("solid", fgColor="FFF3E0"),
        "BASICA":    PatternFill("solid", fgColor="FFF8E1"),
        "INTERMEDIA":PatternFill("solid", fgColor="F3E5F5"),
        "BUENA":     PatternFill("solid", fgColor="E8F5E9"),
        "PROFESIONAL":PatternFill("solid", fgColor="E3F2FD"),
        "ALTA":      PatternFill("solid", fgColor="E8F5E9"),
        "MEDIA":     PatternFill("solid", fgColor="FFF9C4"),
    }

    for i, r in enumerate(resultados, 2):
        ws.row_dimensions[i].height = 72
        row_fill = fills.get(r["calidad_web"], PatternFill("solid", fgColor="FFFFFF"))

        vals = [
            i-1, "", r["nombre"], r["prioridad"], r["sector"], r["municipio"],
            r["empleados"], r["tel_original"], r["telefonos"], r["emails"],
            r["url"], r["calidad_web"], r["chat_ia"], r["whatsapp"],
            r["es_target"], r["facebook"], r["instagram"], r["linkedin"],
            r["youtube"], r["whatsapp_link"], r["descripcion"],
            r["servicios"], r["propuesta"],
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=ci, value=val)
            cell.fill = row_fill
            cell.font = Font(size=9, bold=(ci == 3))
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in (3,8,9,10,21,22,23)))

            if ci == 11 and val:    # URL
                cell.hyperlink = str(val)
                cell.font = Font(color="1565C0", underline="single", size=9)
            if ci in (16,17,18,20) and val:   # redes
                cell.hyperlink = str(val)
                cell.font = Font(color="1565C0", underline="single", size=9)
            if ci == 12:   # calidad web — color propio
                cc = color_cal.get(str(val), "EEEEEE")
                cell.fill = PatternFill("solid", fgColor=cc)
                cell.font = Font(bold=True, size=9,
                                 color="FFFFFF" if val in ("SIN WEB","CAIDA") else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 15:   # TARGET
                cell.fill = PatternFill("solid", fgColor="1B5E20" if val=="SI" else "B71C1C")
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 4:    # prioridad
                cell.fill = fills.get(str(val), row_fill)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Logo
        if r.get("logo") and os.path.exists(r["logo"]):
            try:
                img = XLImage(r["logo"])
                img.width = 65; img.height = 65
                col_l = ws.cell(row=i, column=2).column_letter
                ws.add_image(img, f"{col_l}{i}")
            except Exception:
                pass

    # ── Tab RESUMEN ──
    ws2 = wb.create_sheet(title="RESUMEN", index=0)
    ws2["A1"] = f"INVESTIGACION 100 EMPRESAS - SANTO DOMINGO | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A1"].font = Font(bold=True, size=13, color="0D47A1")

    total    = len(resultados)
    targets  = sum(1 for r in resultados if r["es_target"] == "SI")
    sin_web  = sum(1 for r in resultados if r["calidad_web"] in ("SIN WEB","CAIDA"))
    sin_chat = sum(1 for r in resultados if r["chat_ia"] == "NO")
    con_fb   = sum(1 for r in resultados if r["facebook"])
    con_ig   = sum(1 for r in resultados if r["instagram"])

    stats = [
        ("TOTAL INVESTIGADAS",   total,    ""),
        ("SON TARGET DE VENTA",  targets,  "sin web prof. o sin chat IA"),
        ("SIN SITIO WEB",        sin_web,  "contactar urgente - mayor posibilidad de cierre"),
        ("SIN CHAT IA",          sin_chat, "tienen web pero pierden leads"),
        ("CON FACEBOOK",         con_fb,   "podemos hacer social media + web"),
        ("CON INSTAGRAM",        con_ig,   "audiencia activa - oportunidad de venta"),
    ]
    fills2 = ["E3F2FD","E8F5E9","FFEBEE","FFF9C4","F3E5F5","E8EAF6"]
    for ri, (label, val, nota) in enumerate(stats, 3):
        ws2.cell(row=ri, column=1, value=label).font = Font(bold=True, size=10)
        c = ws2.cell(row=ri, column=2, value=val)
        c.font = Font(bold=True, size=14, color="0D47A1")
        ws2.cell(row=ri, column=3, value=nota).font = Font(size=9, color="616161")
        for col in [1,2,3]:
            ws2.cell(row=ri, column=col).fill = PatternFill("solid", fgColor=fills2[ri-3])

    ws2["A11"] = "RESUMEN POR CALIDAD DE WEB:"
    ws2["A11"].font = Font(bold=True)
    cals = {}
    for r in resultados:
        cals[r["calidad_web"]] = cals.get(r["calidad_web"],0) + 1
    for ri, (cal, cnt) in enumerate(sorted(cals.items(), key=lambda x: x[1], reverse=True), 12):
        ws2.cell(row=ri, column=1, value=cal)
        ws2.cell(row=ri, column=2, value=cnt)
        ws2.cell(row=ri, column=1).fill = PatternFill("solid", fgColor=color_cal.get(cal,"EEEEEE"))

    ws2["A20"] = "PROPUESTAS TOP 10 (por calidad de oportunidad):"
    ws2["A20"].font = Font(bold=True)
    top10 = sorted([r for r in resultados if r["es_target"]=="SI"],
                   key=lambda x: x["score_web"])[:10]
    for ri, r in enumerate(top10, 21):
        ws2.cell(row=ri, column=1, value=r["nombre"]).font = Font(bold=True, size=9)
        ws2.cell(row=ri, column=2, value=r["calidad_web"]).font = Font(size=9)
        ws2.cell(row=ri, column=3, value=r["propuesta"]).font = Font(size=9)

    for col, w in [(1,40),(2,14),(3,72)]:
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = w

    wb.save(OUTPUT)
    return targets, sin_web, sin_chat

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 65)
    print("  INVESTIGADOR 100 EMPRESAS - WEB FACTORY PROSPECT MACHINE")
    print("=" * 65)

    # Cargar progreso previo si existe
    progreso = {}
    if os.path.exists(PROGRESS_F):
        try:
            with open(PROGRESS_F, encoding="utf-8") as f:
                progreso = json.load(f)
            print(f"[+] Progreso anterior encontrado: {len(progreso)} empresas ya investigadas")
        except Exception:
            pass

    print("\nCargando y filtrando empresas...")
    empresas = cargar_empresas()
    print(f"Top {len(empresas)} empresas seleccionadas (15-200 emp, locales, alta prioridad)\n")

    resultados = []
    for idx, emp in enumerate(empresas, 1):
        nombre = emp["nombre"]
        if nombre in progreso:
            print(f"[{idx:3}/{len(empresas)}] SKIP (ya investigada): {nombre}")
            resultados.append(progreso[nombre])
            continue
        try:
            res = investigar(emp, idx, len(empresas))
            resultados.append(res)
            progreso[nombre] = res
            # Guardar progreso cada 10
            if idx % 10 == 0:
                with open(PROGRESS_F, "w", encoding="utf-8") as f:
                    # Convertir None a string para JSON
                    safe = {k: {kk: str(vv) if vv is None else vv
                                for kk, vv in v.items() if kk != "logo"}
                            for k, v in progreso.items()}
                    json.dump(safe, f, ensure_ascii=False, indent=2)
                print(f"\n   [Progreso guardado: {idx}/{len(empresas)}]\n")
        except Exception as e:
            print(f"   ERROR: {e}")
            resultados.append({
                "nombre": nombre, "rnc": emp["rnc"], "municipio": emp["municipio"],
                "sector": emp["sector"], "empleados": emp["empleados"],
                "salarios": emp["salarios"], "prioridad": emp["prioridad"],
                "tel_original": emp["telefono"], "telefonos": emp["telefono"],
                "emails":"","url":"","calidad_web":"ERROR","score_web":0,
                "chat_ia":"?","whatsapp":"?","es_target":"?",
                "facebook":"","instagram":"","linkedin":"","twitter":"",
                "youtube":"","whatsapp_link":"","descripcion":f"Error: {e}",
                "servicios":"","titulo_web":"","logo":None,
                "propuesta":"Revisar manualmente",
            })

    print(f"\n\nExportando Excel con {len(resultados)} empresas...")
    targets, sin_web, sin_chat = exportar(resultados)

    print("\n" + "=" * 65)
    print(f"  COMPLETADO")
    print(f"  Investigadas:    {len(resultados)}")
    print(f"  Targets de venta:{targets}")
    print(f"  Sin sitio web:   {sin_web}")
    print(f"  Sin chat IA:     {sin_chat}")
    print(f"  Archivo:         {OUTPUT}")
    print("=" * 65)
