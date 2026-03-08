"""
Investigador de Empresas - Web Factory
Lee las primeras N empresas del Excel de prospectos,
investiga cada una en la web, y exporta un reporte completo.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import re
import os
import time
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from bs4 import BeautifulSoup
try:
    from ddgs import DDGS
except ImportError:
    from duckduckgo_search import DDGS
from urllib.parse import urljoin, urlparse
from io import BytesIO
from PIL import Image as PILImage
from datetime import datetime

# ─── Configuración ────────────────────────────────────────────────────────────

INPUT_EXCEL  = r"C:\Users\Willy\OneDrive\Escritorio\PROSPECTOS_SD_WILLY.xlsx"
OUTPUT_EXCEL = r"C:\Users\Willy\OneDrive\Escritorio\INVESTIGACION_10_EMPRESAS.xlsx"
LOGOS_DIR    = r"C:\Users\Willy\sistema de egocios\research\logos"
N_EMPRESAS   = 10
HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}

os.makedirs(LOGOS_DIR, exist_ok=True)

# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()

def find_phones(text):
    """Extrae teléfonos del texto (formato RD y otros)."""
    patterns = [
        r"\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}",
        r"\+?1?\s?\(?\d{3}\)?\s?\d{3}[\-\s]\d{4}",
        r"8[0-9]{2}[\-\.\s]?\d{3}[\-\.\s]?\d{4}",
    ]
    phones = []
    for p in patterns:
        phones += re.findall(p, text)
    # Limpiar y deduplicar
    seen = set()
    result = []
    for ph in phones:
        ph = re.sub(r"[^\d\+]", "", ph)
        if ph not in seen and len(ph) >= 7:
            seen.add(ph)
            result.append(ph)
    return result[:5]

def find_emails(text):
    emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return list(set(e for e in emails if "example" not in e and "test" not in e))[:3]

def find_socials(text, url=""):
    combined = (text or "") + " " + (url or "")
    socials = {}
    patterns = {
        "Facebook":  r"facebook\.com/[A-Za-z0-9._\-/]+",
        "Instagram": r"instagram\.com/[A-Za-z0-9._\-/]+",
        "LinkedIn":  r"linkedin\.com/(?:company|in)/[A-Za-z0-9._\-/]+",
        "Twitter/X": r"(?:twitter|x)\.com/[A-Za-z0-9._\-]+",
        "YouTube":   r"youtube\.com/(?:channel|user|@)[A-Za-z0-9._\-/]+",
        "WhatsApp":  r"(?:wa\.me|whatsapp\.com)/[0-9]+",
    }
    for red, pat in patterns.items():
        m = re.search(pat, combined, re.IGNORECASE)
        if m:
            socials[red] = "https://" + m.group(0).lstrip("/")
    return socials

def tiene_chat_ia(soup, html):
    keywords = ["tawk", "intercom", "zendesk", "drift", "hubspot", "tidio",
                 "crisp", "freshchat", "chatbot", "live chat", "chat en vivo",
                 "asistente virtual", "whatsapp", "wa.me"]
    html_lower = html.lower()
    return any(kw in html_lower for kw in keywords)

def tiene_whatsapp_form(soup, html):
    wa_keywords = ["whatsapp", "wa.me", "api.whatsapp"]
    return any(kw in html.lower() for kw in wa_keywords)

def calificar_web(soup, html, url):
    if not url:
        return "SIN WEB", 0
    score = 0
    notas = []
    if soup:
        score += 1
        if soup.find("meta", {"name": "description"}): score += 1
        imgs = soup.find_all("img")
        if len(imgs) > 3: score += 1
        if soup.find("form"): score += 1; notas.append("tiene formulario")
        if tiene_chat_ia(soup, html): score += 2; notas.append("chat/IA detectado")
        if tiene_whatsapp_form(soup, html): score += 1; notas.append("WhatsApp")
        if soup.find("meta", {"name": "viewport"}): score += 1; notas.append("responsive")
    calidad = {0: "SIN WEB", 1: "BASICA", 2: "BASICA", 3: "INTERMEDIA",
               4: "INTERMEDIA", 5: "BUENA", 6: "BUENA", 7: "PROFESIONAL", 8: "PROFESIONAL"}
    return calidad.get(min(score, 8), "BASICA"), score

def descargar_logo(url_base, nombre_empresa, soup):
    """Intenta descargar favicon u OG image."""
    if not soup or not url_base:
        return None
    try:
        # Buscar OG image primero
        og = soup.find("meta", property="og:image") or soup.find("meta", attrs={"name": "og:image"})
        img_url = None
        if og and og.get("content"):
            img_url = urljoin(url_base, og["content"])
        else:
            # Favicon
            fav = soup.find("link", rel=lambda r: r and "icon" in " ".join(r).lower())
            if fav and fav.get("href"):
                img_url = urljoin(url_base, fav["href"])
            else:
                img_url = urljoin(url_base, "/favicon.ico")

        r = requests.get(img_url, timeout=8, headers=HEADERS_HTTP)
        if r.status_code == 200 and len(r.content) > 500:
            img = PILImage.open(BytesIO(r.content)).convert("RGBA")
            img = img.resize((80, 80), PILImage.LANCZOS)
            slug = re.sub(r"[^\w]", "_", nombre_empresa[:20])
            path = os.path.join(LOGOS_DIR, f"{slug}.png")
            img.save(path, "PNG")
            return path
    except Exception:
        pass
    return None

# ─── Búsqueda web principal ───────────────────────────────────────────────────

def buscar_empresa(nombre, rnc=None):
    """Busca la empresa con DuckDuckGo — multiples queries para mejor cobertura."""
    # Simplificar nombre (quitar SA, SRL, EIRL, etc.)
    nombre_limpio = re.sub(r"\b(S\.?\s*A\.?S?|SRL|EIRL|LTD|INC|CORP|S\.?\s*A\.?|CIA|COMPANIA|COMPANY)\b", "", nombre, flags=re.IGNORECASE).strip()
    nombre_limpio = re.sub(r"\s+", " ", nombre_limpio).strip()

    queries = [
        f'"{nombre_limpio}" "república dominicana" OR "santo domingo" web site',
        f'{nombre_limpio} dominicana contacto telefono',
        f'site:*.do {nombre_limpio}',
    ]
    if rnc:
        queries.insert(0, f'RNC {rnc} {nombre_limpio}')

    all_results = []
    for query in queries[:2]:  # Solo 2 queries para no saturar
        try:
            with DDGS() as ddg:
                res = list(ddg.text(query, max_results=6, region="es-419"))
                all_results.extend(res)
            time.sleep(0.8)
        except Exception:
            pass
    return all_results

def scrape_sitio(url):
    """Descarga y parsea el sitio web de la empresa."""
    if not url:
        return None, "", url
    try:
        parsed = urlparse(url)
        if not parsed.scheme:
            url = "https://" + url
        r = requests.get(url, timeout=12, headers=HEADERS_HTTP, allow_redirects=True)
        r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, "lxml")
        return soup, r.text, r.url
    except Exception as e:
        try:
            url_http = url.replace("https://", "http://")
            r = requests.get(url_http, timeout=10, headers=HEADERS_HTTP, allow_redirects=True)
            soup = BeautifulSoup(r.text, "lxml")
            return soup, r.text, r.url
        except Exception:
            return None, "", url

def extraer_info_web(soup, html, url_final):
    """Extrae toda la info útil del HTML."""
    info = {
        "descripcion": "",
        "telefono_web": "",
        "email_web": "",
        "servicios": "",
        "direccion_web": "",
        "redes": {},
        "titulo": "",
    }
    if not soup:
        return info

    # Título
    title = soup.find("title")
    if title:
        info["titulo"] = clean(title.get_text())

    # Meta description
    meta = soup.find("meta", {"name": "description"}) or soup.find("meta", property="og:description")
    if meta:
        info["descripcion"] = clean(meta.get("content", ""))[:300]

    # Texto completo para extracción
    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)

    # Teléfonos
    phones = find_phones(text) + find_phones(html)
    info["telefono_web"] = " | ".join(list(dict.fromkeys(phones))[:3])

    # Emails
    emails = find_emails(text) + find_emails(html)
    info["email_web"] = " | ".join(list(dict.fromkeys(emails))[:2])

    # Redes sociales
    info["redes"] = find_socials(html, url_final)

    # Servicios - párrafos relevantes
    servicios_kw = ["servicio", "producto", "ofrecemos", "soluciones", "especializ",
                    "brindamos", "nuestros", "hacemos", "realizamos"]
    parrafos = soup.find_all(["p", "li", "h2", "h3"], limit=50)
    servicios = []
    for p in parrafos:
        t = clean(p.get_text())
        if len(t) > 20 and any(kw in t.lower() for kw in servicios_kw):
            servicios.append(t[:120])
        if len(servicios) >= 5:
            break
    info["servicios"] = " | ".join(servicios) if servicios else ""

    return info

def detectar_url_en_ddg(results, nombre):
    """Encuentra la URL del sitio OFICIAL en resultados DDG con scoring."""
    dominios_excluir = [
        "facebook.com", "instagram.com", "linkedin.com", "twitter.com", "x.com",
        "youtube.com", "google.com", "yelp.com", "paginas.amarillas", "yellowpages",
        "wikipedia.org", "wa.me", "whatsapp.com", "reddit.com", "quora.com",
        "tripadvisor.com", "trustpilot.com", "g2.com", "glassdoor.com",
        "zhihu.com", "baidu.com", "chatgpt.com", "drudgereport.com", "xfinity.com",
        "wordreference.com", "amazon.com", "ebay.com", "mercadolibre.com",
        "rainews.it", "digitei.com", "gulfcoast", "forums.",
    ]

    # Palabras del nombre de la empresa (ignorar stopwords)
    stopwords = {"sa", "sas", "srl", "eirl", "ltd", "inc", "corp", "de", "la", "el",
                 "los", "las", "del", "y", "e", "the", "&", "dominicana", "dominicano",
                 "republica", "company", "compania", "cia", "hermanos", "interamerica"}
    palabras = [w.lower() for w in re.split(r"[\s\-_&]+", nombre)
                if len(w) >= 4 and w.lower() not in stopwords]

    scored = []
    for r in results:
        url = r.get("href", "")
        title = (r.get("title", "") or "").lower()
        body = (r.get("body", "") or "").lower()

        if not url:
            continue
        # Excluir dominios basura
        if any(d in url.lower() for d in dominios_excluir):
            continue

        score = 0
        url_lower = url.lower()

        # Bonus si la URL contiene parte del nombre
        for pal in palabras:
            if pal in url_lower:
                score += 3
            if pal in title:
                score += 2
            if pal in body:
                score += 1

        # Bonus si es dominio .do (dominicano)
        if ".do/" in url_lower or url_lower.endswith(".do"):
            score += 4

        # Bonus si el body menciona RD
        if any(kw in body for kw in ["santo domingo", "republica dominicana", "rd", "dominicana"]):
            score += 2

        # Penalizar si parece directorio/noticias
        if any(kw in url_lower for kw in ["/noticias/", "/news/", "/articulos/", "/blog/", "/foro/", "/forum/"]):
            score -= 3

        if score > 0:
            scored.append((score, url))

    if scored:
        scored.sort(key=lambda x: x[0], reverse=True)
        return scored[0][1]

    return ""  # No encontrado con confianza

# ─── Proceso principal ────────────────────────────────────────────────────────

def investigar_empresa(empresa, idx):
    nombre = empresa["nombre"]
    rnc = empresa["rnc"]
    tel_excel = empresa["telefono"]

    print(f"\n[{idx}/10] Investigando: {nombre}")
    print(f"         RNC: {rnc} | Tel Excel: {tel_excel}")

    # 1. Búsqueda DDG
    print("         -> Buscando en web...")
    results = buscar_empresa(nombre, rnc)
    time.sleep(1.5)  # respetar rate limit

    # 2. Snippets DDG para info rápida
    snippet_text = " ".join(r.get("body", "") for r in results)
    phones_ddg = find_phones(snippet_text)
    emails_ddg = find_emails(snippet_text)
    socials_ddg = find_socials(snippet_text)

    # 3. Detectar URL oficial
    url_oficial = detectar_url_en_ddg(results, nombre)
    print(f"         -> Web encontrada: {url_oficial or 'NO ENCONTRADA'}")

    # 4. Scrape del sitio
    soup, html, url_final = None, "", url_oficial
    logo_path = None

    if url_oficial:
        print("         -> Analizando sitio web...")
        soup, html, url_final = scrape_sitio(url_oficial)
        time.sleep(1)

        if soup:
            print("         -> Descargando imagen/logo...")
            logo_path = descargar_logo(url_final, nombre, soup)

    # 5. Extraer info del sitio
    info = extraer_info_web(soup, html, url_final)

    # 6. Combinar info DDG + sitio
    todos_phones = list(dict.fromkeys(
        ([tel_excel] if tel_excel else []) +
        (info["telefono_web"].split(" | ") if info["telefono_web"] else []) +
        phones_ddg
    ))
    todos_emails = list(dict.fromkeys(
        (info["email_web"].split(" | ") if info["email_web"] else []) +
        emails_ddg
    ))
    todas_redes = {**socials_ddg, **info["redes"]}

    # 7. Calificar web
    calidad_web, score_web = calificar_web(soup, html, url_oficial)
    tiene_chat = tiene_chat_ia(soup, html) if soup else False
    tiene_wa = tiene_whatsapp_form(soup, html) if soup else False

    # 8. Descripción del negocio
    desc = info["descripcion"]
    if not desc and results:
        desc = clean(results[0].get("body", ""))[:300]

    resultado = {
        "nombre": nombre,
        "rnc": rnc,
        "municipio": empresa["municipio"],
        "sector": empresa["sector"],
        "empleados": empresa["empleados"],
        "salarios": empresa["salarios_mensual"],
        "tel_original": tel_excel,
        "todos_telefonos": " | ".join(t for t in todos_phones if t)[:100],
        "emails": " | ".join(todos_emails)[:100],
        "url_oficial": url_final or url_oficial,
        "calidad_web": calidad_web,
        "score_web": score_web,
        "tiene_chat_ia": "SI" if tiene_chat else "NO",
        "tiene_whatsapp": "SI" if tiene_wa else "NO",
        "es_target": "NO" if (calidad_web in ("PROFESIONAL", "BUENA") and tiene_chat) else "SI",
        "facebook": todas_redes.get("Facebook", ""),
        "instagram": todas_redes.get("Instagram", ""),
        "linkedin": todas_redes.get("LinkedIn", ""),
        "twitter": todas_redes.get("Twitter/X", ""),
        "whatsapp_link": todas_redes.get("WhatsApp", ""),
        "descripcion": desc,
        "servicios": info["servicios"][:400] if info["servicios"] else "",
        "titulo_web": info["titulo"],
        "logo_path": logo_path,
        "propuesta": generar_propuesta_rapida(nombre, calidad_web, tiene_chat, tiene_wa, empresa["sector"]),
    }

    print(f"         OK Web: {calidad_web} | Chat IA: {resultado['tiene_chat_ia']} | Target: {resultado['es_target']}")
    return resultado

def generar_propuesta_rapida(nombre, calidad_web, tiene_chat, tiene_wa, sector):
    """Genera un argumento de venta de 2 líneas."""
    if calidad_web == "SIN WEB":
        return f"{nombre} no tiene presencia web. Podemos lanzar su sitio profesional con chat IA y captura por WhatsApp en 48 horas."
    elif calidad_web in ("BASICA", "INTERMEDIA"):
        return f"El sitio de {nombre} es básico y sin automatización. Un rediseño con IA y WhatsApp puede multiplicar sus consultas."
    elif not tiene_chat:
        return f"{nombre} tiene web pero sin asistente IA ni captura por WhatsApp. Están perdiendo leads cada noche."
    else:
        return f"{nombre} tiene buena presencia digital. Explorar mejoras de conversión o servicios adicionales."

# ─── Leer prospectos del Excel ────────────────────────────────────────────────

def leer_prospectos(n=10):
    wb = openpyxl.load_workbook(INPUT_EXCEL, read_only=True, data_only=True)
    # Buscar hoja ALTA
    hoja = None
    for name in wb.sheetnames:
        if "ALTA" in name.upper():
            hoja = wb[name]
            break
    if not hoja:
        hoja = wb.active

    empresas = []
    headers = None
    for row in hoja.iter_rows(min_row=1, values_only=True):
        if headers is None:
            headers = [str(c).upper().strip() if c else "" for c in row]
            continue
        if not row[0]:
            continue
        def col(name_keys):
            for k in name_keys:
                if k in headers:
                    v = row[headers.index(k)]
                    return v if v is not None else ""
            return ""

        empresas.append({
            "rnc":             col(["RNC"]),
            "nombre":          str(col(["EMPRESA"]) or "").strip(),
            "empleados":       col(["EMPLEADOS"]) or 0,
            "salarios_mensual":col(["SALARIO MASA"]) or 0,
            "municipio":       str(col(["MUNICIPIO"]) or ""),
            "sector":          str(col(["SECTOR"]) or ""),
            "telefono":        str(col(["TELEFONO"]) or ""),
            "direccion":       str(col(["DIRECCION"]) or ""),
        })
        if len(empresas) >= n:
            break

    wb.close()
    return empresas

# ─── Exportar Excel ───────────────────────────────────────────────────────────

def exportar_excel(resultados, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INVESTIGACION"

    # Paleta
    H_FILL  = PatternFill("solid", fgColor="0D47A1")
    G_FILL  = PatternFill("solid", fgColor="E8F5E9")
    W_FILL  = PatternFill("solid", fgColor="FFFFFF")
    R_FILL  = PatternFill("solid", fgColor="FFEBEE")
    Y_FILL  = PatternFill("solid", fgColor="FFF9C4")
    G2_FILL = PatternFill("solid", fgColor="E3F2FD")

    COLS = [
        ("#",                3),  ("LOGO",            10),
        ("EMPRESA",         38),  ("SECTOR",          20),
        ("MUNICIPIO",       18),  ("EMPLEADOS",        10),
        ("TEL ORIGINAL",    15),  ("TODOS TELEFONOS",  25),
        ("EMAILS",          28),  ("SITIO WEB",        32),
        ("CALIDAD WEB",     14),  ("CHAT IA",          10),
        ("WHATSAPP",        10),  ("ES TARGET",        10),
        ("FACEBOOK",        30),  ("INSTAGRAM",        30),
        ("LINKEDIN",        30),  ("DESCRIPCION",      50),
        ("SERVICIOS",       55),  ("PROPUESTA VENTA",  60),
    ]

    # Encabezados
    ws.row_dimensions[1].height = 35
    for col_idx, (header, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = H_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "C2"

    # Filas de datos
    for i, r in enumerate(resultados, 2):
        ws.row_dimensions[i].height = 75

        # Color de fila según target
        if r["es_target"] == "SI" and r["calidad_web"] in ("SIN WEB", "BASICA"):
            row_fill = G_FILL   # verde claro = oportunidad directa
        elif r["es_target"] == "SI":
            row_fill = Y_FILL   # amarillo = oportunidad con web parcial
        else:
            row_fill = R_FILL   # rojo = ya tienen todo, no es target

        datos_fila = [
            i - 1,
            "",  # logo (se inserta imagen)
            r["nombre"],
            r["sector"],
            r["municipio"],
            r["empleados"],
            r["tel_original"],
            r["todos_telefonos"],
            r["emails"],
            r["url_oficial"],
            r["calidad_web"],
            r["tiene_chat_ia"],
            r["tiene_whatsapp"],
            r["es_target"],
            r["facebook"],
            r["instagram"],
            r["linkedin"],
            r["descripcion"],
            r["servicios"],
            r["propuesta"],
        ]

        for col_idx, val in enumerate(datos_fila, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=9, bold=(col_idx == 3))

            # Colorear columna calidad web
            if col_idx == 11:
                color = {
                    "SIN WEB": "EF5350", "BASICA": "FF8A65",
                    "INTERMEDIA": "FFD54F", "BUENA": "81C784", "PROFESIONAL": "42A5F5"
                }.get(str(val), "EEEEEE")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, size=9)

            # ES TARGET: resaltar
            if col_idx == 14:
                cell.fill = PatternFill("solid", fgColor="1B5E20" if val == "SI" else "B71C1C")
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # URL como link
            if col_idx == 10 and val:
                cell.hyperlink = str(val)
                cell.font = Font(color="1565C0", underline="single", size=9)

            # Redes como links
            if col_idx in (15, 16, 17) and val:
                cell.hyperlink = str(val)
                cell.font = Font(color="1565C0", underline="single", size=9)

        # Insertar logo si existe
        if r.get("logo_path") and os.path.exists(r["logo_path"]):
            try:
                img = XLImage(r["logo_path"])
                img.width = 70
                img.height = 70
                col_b = ws.cell(row=i, column=2).column_letter
                ws.add_image(img, f"{col_b}{i}")
            except Exception:
                pass

    # Hoja resumen rápido
    ws2 = wb.create_sheet(title="RESUMEN")
    ws2["A1"] = f"INVESTIGACION - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A1"].font = Font(bold=True, size=14, color="0D47A1")

    total = len(resultados)
    targets = sum(1 for r in resultados if r["es_target"] == "SI")
    sin_web = sum(1 for r in resultados if r["calidad_web"] == "SIN WEB")
    sin_chat = sum(1 for r in resultados if r["tiene_chat_ia"] == "NO")

    ws2["A3"] = "EMPRESAS INVESTIGADAS"
    ws2["B3"] = total
    ws2["A4"] = "SON TARGET (sin web prof. o sin IA)"
    ws2["B4"] = targets
    ws2["A4"].font = Font(bold=True, color="1B5E20")
    ws2["B4"].font = Font(bold=True, size=14, color="1B5E20")
    ws2["A5"] = "SIN SITIO WEB"
    ws2["B5"] = sin_web
    ws2["A6"] = "SIN CHAT IA"
    ws2["B6"] = sin_chat
    ws2["A8"] = "PROPUESTA RECOMENDADA POR EMPRESA:"
    ws2["A8"].font = Font(bold=True)
    for i, r in enumerate(resultados, 9):
        ws2[f"A{i}"] = r["nombre"]
        ws2[f"B{i}"] = r["calidad_web"]
        ws2[f"C{i}"] = r["propuesta"]
        ws2[f"A{i}"].font = Font(bold=True, size=9)
        ws2[f"C{i}"].font = Font(size=9)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 70

    wb.save(output_path)
    print(f"\n[OK] Excel guardado: {output_path}")
    return targets, sin_web

# ─── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  INVESTIGADOR DE EMPRESAS - WEB FACTORY")
    print("=" * 60)
    print(f"\nLeyendo primeras {N_EMPRESAS} empresas de ALTA prioridad...")

    prospectos = leer_prospectos(N_EMPRESAS)
    print(f"Empresas cargadas: {len(prospectos)}")
    for p in prospectos:
        print(f"  - {p['nombre']} ({p['empleados']} emp) | {p['sector']}")

    print(f"\nIniciando investigación web...")
    resultados = []
    for idx, empresa in enumerate(prospectos, 1):
        try:
            res = investigar_empresa(empresa, idx)
            resultados.append(res)
        except Exception as e:
            print(f"   [ERROR] {empresa['nombre']}: {e}")
            resultados.append({
                "nombre": empresa["nombre"], "rnc": empresa["rnc"],
                "municipio": empresa["municipio"], "sector": empresa["sector"],
                "empleados": empresa["empleados"], "salarios": 0,
                "tel_original": empresa["telefono"], "todos_telefonos": empresa["telefono"],
                "emails": "", "url_oficial": "", "calidad_web": "ERROR",
                "score_web": 0, "tiene_chat_ia": "?", "tiene_whatsapp": "?",
                "es_target": "?", "facebook": "", "instagram": "", "linkedin": "",
                "twitter": "", "whatsapp_link": "", "descripcion": f"Error: {e}",
                "servicios": "", "titulo_web": "", "logo_path": None,
                "propuesta": "",
            })

    print(f"\nExportando Excel...")
    targets, sin_web = exportar_excel(resultados, OUTPUT_EXCEL)

    print("\n" + "=" * 60)
    print(f"  INVESTIGACION COMPLETADA")
    print(f"  Empresas investigadas: {len(resultados)}")
    print(f"  Son target de venta:   {targets}")
    print(f"  Sin sitio web:         {sin_web}")
    print(f"  Archivo:               {OUTPUT_EXCEL}")
    print("=" * 60)
