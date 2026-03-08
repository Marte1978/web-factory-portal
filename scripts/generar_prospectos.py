"""
Genera y segmenta prospectos desde el Excel de 69,313 empresas.
Puede subir directamente a Supabase en batches.

Uso:
  py scripts/generar_prospectos.py                        # genera Excel local
  py scripts/generar_prospectos.py --upload               # sube batch 100 a Supabase
  py scripts/generar_prospectos.py --upload --batch 50    # batch de 50
  py scripts/generar_prospectos.py --upload --offset 100  # empieza desde empresa 100
  py scripts/generar_prospectos.py --upload --sector gold --batch 200
  py scripts/generar_prospectos.py --upload --municipio "SANTIAGO"
  py scripts/generar_prospectos.py --stats               # solo muestra estadisticas
"""
import sys, io, os, re, json, time, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import requests

# ─── Config ───────────────────────────────────────────────────────────────────
EXCEL_IN  = r"C:\Users\Willy\Downloads\empresas_2.xlsx"
EXCEL_OUT = r"C:\Users\Willy\OneDrive\Escritorio\PROSPECTOS_SD_WILLY.xlsx"

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "https://fmjvktaaxsdhukbkefnw.supabase.co")
SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZtanZrdGFheHNkaHVrYmtlZm53Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3Mjk5MTMyNCwiZXhwIjoyMDg4NTY3MzI0fQ.bcRkFjlI52ZmOtFSeIaaqB84qV9Tk_Cf1SiMnVm7sBE")

SUPABASE_HEADERS = {
    "apikey":        SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=minimal,resolution=merge-duplicates",
}

# ─── Sector maps ──────────────────────────────────────────────────────────────
SECTORES_GOLD = {8,11,12,14,15,16,18,7,10}
SECTORES_SILVER = {3,5,6,9,33,41}
SECTORES_EXCLUIR = {1,2,4,13,17,25,26,34,36}

SECTOR_NOMBRE = {
    3:'Manufactura', 5:'Construccion', 6:'Comercio Mayor',
    7:'Comercio Minorista', 8:'Hoteles/Turismo', 9:'Transporte/Logistica',
    10:'Finanzas/Bancos', 11:'Inmobiliaria', 12:'Servicios Empresariales',
    14:'Educacion', 15:'Salud/Medicina', 16:'Servicios Sociales',
    18:'Otros Servicios', 33:'Manufactura Avanzada', 41:'Construccion Especializada',
}

MUNICIPIOS_SD = {
    'DISTRITO NACIONAL','SANTO DOMINGO ESTE','SANTO DOMINGO OESTE','SANTO DOMINGO NORTE'
}

MULTINACIONAL_KW = [
    'colgate','palmolive','kpmg','pricewaterhouse','deloitte','ernst&young',
    'ernst & young','bayer','kimberly','mondelez','mead johnson','3m dominicana',
    'american airlines','scotiabank','citibank','honda','toyota','ford ',
    'hyundai','samsung','apple ','microsoft','oracle ','ibm ','xerox',
    'siemens','nestle','unilever','procter','johnson & johnson','pfizer',
    'novartis','roche ','dominican power',
]


def slug(nombre: str) -> str:
    return re.sub(r"[^\w]", "_", nombre.lower())[:35]


def es_multinacional(nombre: str) -> bool:
    n = nombre.lower()
    return any(kw in n for kw in MULTINACIONAL_KW)


def parse_args():
    p = argparse.ArgumentParser(description="Genera prospectos desde el Excel de 69K empresas")
    p.add_argument("--upload",     action="store_true", help="Subir a Supabase")
    p.add_argument("--stats",      action="store_true", help="Solo mostrar estadisticas")
    p.add_argument("--batch",      type=int,   default=100,  help="Empresas por batch (default: 100)")
    p.add_argument("--offset",     type=int,   default=0,    help="Saltar primeras N empresas")
    p.add_argument("--sector",     type=str,   default="gold",
                   choices=["gold","silver","all"], help="Filtrar por tipo de sector")
    p.add_argument("--municipio",  type=str,   default=None,
                   help="Municipio(s) separados por coma. Default: los 4 de Santo Domingo")
    p.add_argument("--emp-min",    type=int,   default=15,   help="Empleados minimo")
    p.add_argument("--emp-max",    type=int,   default=500,  help="Empleados maximo")
    p.add_argument("--excluir-mult", action="store_true", default=True,
                   help="Excluir multinacionales")
    p.add_argument("--output",     type=str,   default=EXCEL_OUT, help="Ruta del Excel de salida")
    return p.parse_args()


def cargar_empresas(args) -> list:
    """Lee el Excel y aplica filtros."""
    if not os.path.exists(EXCEL_IN):
        print(f"ERROR: No se encontro el archivo Excel en:\n  {EXCEL_IN}")
        sys.exit(1)

    # Municipios a filtrar
    if args.municipio:
        munis = {m.strip().upper() for m in args.municipio.split(",")}
    else:
        munis = MUNICIPIOS_SD

    print(f"Leyendo {EXCEL_IN}...")
    wb = openpyxl.load_workbook(EXCEL_IN, read_only=True, data_only=True)
    ws = wb.active

    empresas = []
    omitidas = {"sector_excluido": 0, "municipio": 0, "empleados": 0, "multinacional": 0}
    total_leidas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 16:
            continue
        rnc, nombre, calle, num, edif, piso, apto, barrio, muni_cod, muni_nom, \
            tipo, tel1, ext1, emp, sal, sec_eco = row[:16]

        total_leidas += 1
        muni = str(muni_nom or "").strip().upper()
        nombre_s = str(nombre or "").strip()

        if not nombre_s:
            continue
        if muni not in munis:
            omitidas["municipio"] += 1
            continue
        if not isinstance(emp, (int, float)) or emp < args.emp_min or emp > args.emp_max:
            omitidas["empleados"] += 1
            continue

        sec = int(sec_eco) if isinstance(sec_eco, (int, float)) else 0
        if sec in SECTORES_EXCLUIR:
            omitidas["sector_excluido"] += 1
            continue

        if args.sector == "gold" and sec not in SECTORES_GOLD:
            omitidas["sector_excluido"] += 1
            continue
        if args.sector == "silver" and sec not in SECTORES_SILVER:
            omitidas["sector_excluido"] += 1
            continue
        if args.sector == "all" and sec not in SECTORES_GOLD and sec not in SECTORES_SILVER:
            omitidas["sector_excluido"] += 1
            continue

        if args.excluir_mult and es_multinacional(nombre_s):
            omitidas["multinacional"] += 1
            continue

        is_gold = sec in SECTORES_GOLD
        prio_n  = 3 if is_gold else 2
        sal_v   = sal if isinstance(sal, (int, float)) else 0
        score   = (emp * prio_n) + (sal_v / 1200)

        tel_s = ""
        if isinstance(tel1, (int, float)):
            t = str(int(tel1))
            tel_s = t if len(t) >= 7 else ""
        elif tel1:
            tel_s = str(tel1).strip()

        dir_parts = [str(calle or ""), str(num or ""), str(barrio or "")]
        direc = ", ".join(p for p in dir_parts if p and p not in ("None",""))

        empresas.append({
            "id":        slug(nombre_s),
            "rnc":       str(rnc or ""),
            "nombre":    nombre_s,
            "empleados": int(emp),
            "salarios":  sal_v,
            "municipio": muni,
            "sector":    SECTOR_NOMBRE.get(sec, f"Sector-{sec}"),
            "prioridad": "ALTA" if is_gold else "MEDIA",
            "score":     round(score, 2),
            "telefono":  tel_s,
            "direccion": direc,
        })

    wb.close()
    empresas.sort(key=lambda x: x["score"], reverse=True)

    print(f"\nTotal leidas:      {total_leidas:,}")
    print(f"Filtradas (pasan): {len(empresas):,}")
    print(f"  Municipio:       {omitidas['municipio']:,} excluidas")
    print(f"  Empleados:       {omitidas['empleados']:,} fuera de rango")
    print(f"  Sector:          {omitidas['sector_excluido']:,} excluidas")
    print(f"  Multinacional:   {omitidas['multinacional']:,} excluidas")

    return empresas


def upload_batch(empresas: list, batch: list, offset: int, batch_size: int) -> int:
    """Sube un batch de empresas a Supabase. Retorna cantidad subida."""
    print(f"\nSubiendo batch {offset+1}–{offset+len(batch)} a Supabase...")

    # Check which already exist
    ids = [c["id"] for c in batch]
    exist_r = requests.get(
        f"{SUPABASE_URL}/rest/v1/companies?id=in.({','.join(ids)})&select=id",
        headers=SUPABASE_HEADERS, timeout=15
    )
    existing = {r["id"] for r in (exist_r.json() if exist_r.status_code == 200 else [])}

    ok = 0
    for c in batch:
        # Only upload columns that exist in the schema
        row = {
            "id": c["id"], "nombre": c["nombre"], "sector": c["sector"],
            "municipio": c["municipio"], "empleados": c["empleados"],
            "telefono": c["telefono"], "direccion": c["direccion"],
            "prioridad": c["prioridad"], "score": c["score"], "rnc": c["rnc"],
            "researched": False, "package_ready": False,
        }
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/companies",
            headers=SUPABASE_HEADERS,
            json=row, timeout=15
        )
        if r.status_code in (200, 201):
            ok += 1
            status = "SKIP (ya existe)" if c["id"] in existing else "OK"
            print(f"  [{ok:3}] {c['nombre'][:40]:40} {status}")
        else:
            print(f"  ERR {c['nombre'][:35]}: {r.text[:80]}")
        time.sleep(0.04)

    return ok


def exportar_excel(empresas: list, output: str):
    """Exporta el Excel con tabs ALTA/MEDIA + RESUMEN."""
    alta  = [e for e in empresas if e["prioridad"] == "ALTA"]
    media = [e for e in empresas if e["prioridad"] == "MEDIA"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADERS = ["#","RNC","EMPRESA","EMPLEADOS","SECTOR","MUNICIPIO",
               "TELEFONO","DIRECCION","SALARIO MASA","ESTADO WEB","CONTACTO","NOTAS"]
    WIDTHS  = [5,13,42,11,22,20,14,35,14,18,20,25]

    def make_tab(tab_name, datos, h_hex, r_hex):
        ws = wb.create_sheet(title=tab_name)
        hfill = PatternFill("solid", fgColor=h_hex)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = Font(color="FFFFFF", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        alt_fill   = PatternFill("solid", fgColor=r_hex)
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        for i, e in enumerate(datos, 1):
            r = i + 1
            row_data = [i, e["rnc"], e["nombre"], e["empleados"], e["sector"],
                        e["municipio"], e["telefono"], e["direccion"], e["salarios"],"","",""]
            fill = alt_fill if i % 2 == 0 else white_fill
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.font = Font(bold=(col == 3), size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=(col in (3, 8)))
        for col, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = "C2"

    make_tab(f"ALTA ({len(alta)})",  alta,  "1B5E20", "E8F5E9")
    make_tab(f"MEDIA ({len(media)})", media, "E65100", "FFF3E0")

    ws_r = wb.create_sheet(title="RESUMEN", index=0)
    ws_r["A1"] = f"PIPELINE DE PROSPECCION — {len(empresas):,} empresas"
    ws_r["A1"].font = Font(bold=True, size=14, color="1B5E20")
    for ri, (label, cnt) in enumerate([
        ("ALTA prioridad", len(alta)), ("MEDIA prioridad", len(media)), ("TOTAL", len(empresas))
    ], 3):
        ws_r.cell(row=ri, column=1, value=label).font = Font(bold=True)
        ws_r.cell(row=ri, column=2, value=cnt).font = Font(bold=True, size=13, color="1B5E20")
    ws_r.column_dimensions["A"].width = 35
    ws_r.column_dimensions["B"].width = 12

    wb.save(output)
    print(f"\nExcel guardado en: {output}")


def main():
    args = parse_args()

    print("=" * 60)
    print("  GENERADOR DE PROSPECTOS — Web Factory Pipeline")
    print("=" * 60)
    print(f"Sector: {args.sector.upper()} | Municipio: {args.municipio or 'Santo Domingo (4)'}",
          f"\nEmpleados: {args.emp_min}–{args.emp_max} | Batch: {args.batch} | Offset: {args.offset}")

    empresas = cargar_empresas(args)

    if args.stats:
        alta  = sum(1 for e in empresas if e["prioridad"] == "ALTA")
        media = sum(1 for e in empresas if e["prioridad"] == "MEDIA")
        print(f"\nRESUMEN: {len(empresas):,} total | {alta:,} ALTA | {media:,} MEDIA")
        print(f"Batches disponibles: {len(empresas) // args.batch} batches de {args.batch}")
        return

    if args.upload:
        # Get batch slice
        batch = empresas[args.offset : args.offset + args.batch]
        if not batch:
            print(f"\nNo hay empresas en offset {args.offset}. Total filtradas: {len(empresas)}")
            return

        ok = upload_batch(empresas, batch, args.offset, args.batch)
        next_offset = args.offset + args.batch

        print(f"\n{'='*60}")
        print(f"  BATCH COMPLETADO")
        print(f"  Subidas: {ok}/{len(batch)}")
        print(f"  Para continuar con el siguiente batch:")
        print(f"  py scripts/generar_prospectos.py --upload --offset {next_offset} --batch {args.batch}")
        if next_offset < len(empresas):
            restantes = len(empresas) - next_offset
            print(f"  Quedan: {restantes:,} empresas ({restantes // args.batch} batches mas)")
        print(f"{'='*60}")
    else:
        exportar_excel(empresas, args.output)
        alta  = sum(1 for e in empresas if e["prioridad"] == "ALTA")
        media = sum(1 for e in empresas if e["prioridad"] == "MEDIA")
        print(f"\nRESUMEN: {len(empresas):,} | ALTA: {alta:,} | MEDIA: {media:,}")
        print("Para subir el primer batch a Supabase:")
        print(f"  py scripts/generar_prospectos.py --upload --batch {args.batch}")


if __name__ == "__main__":
    main()
